import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render


import os
import openpyxl
from django.conf import settings
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime, timedelta
from django.db.models import Count
from .models import Robot
from django.http import HttpResponse


@method_decorator(csrf_exempt, name='dispatch')
class RobotAPIView(View):
    def post(self, request):
        try:
            # Парсинг входных данных
            data = json.loads(request.body)
            serial = data.get('serial')
            model = data.get('model')
            version = data.get('version')
            created = datetime.strptime(data.get('created'), '%Y-%m-%d %H:%M:%S')

            # Проверка на существование модели
            if not Robot.objects.filter(model=model).exists():
                return JsonResponse({'error': 'Invalid model.'}, status=400)


            # Проверка на валидность данных (в данном примере просто проверяем длину)
            if len(serial) > 5 or len(model) > 2 or len(version) > 2:
                return JsonResponse({'error': 'Invalid data. Exceeded maximum length.'}, status=400)

            # Сохранение данных в базу данных
            robot = Robot.objects.create(serial=serial, model=model, version=version, created=created)
            robot.save()

            return JsonResponse({'message': 'Robot information saved successfully.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data.'}, status=400)


def generate_excel_summary(request):
    # Получение текущей даты
    today = datetime.today().strftime('%Y-%m-%d')

    # Формирование пути к файлу Excel
    file_name = f'summary_{today}.xlsx'
    excel_file_path = os.path.join(settings.STATIC_ROOT, file_name)

    # Получение данных о количестве роботов для каждой модели и версии за последнюю неделю
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    robot_data = Robot.objects.filter(created__range=(start_date, end_date)) \
        .values('model', 'version') \
        .annotate(quantity=Count('id'))

    # Создание Excel-файла
    wb = openpyxl.Workbook()

    # Создание страниц для каждой модели
    unique_models = robot_data.values_list('model', flat=True).distinct()
    for model in unique_models:
        ws = wb.create_sheet(title=model)
        df = pd.DataFrame(robot_data.filter(model=model))
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Удаление изначальной пустой страницы
    del wb[wb.sheetnames[0]]

    # Сохранение файла
    wb.save(excel_file_path)

    # Открытие и чтение файла в бинарном режиме
    with open(excel_file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        return response




def index_view(request):
    return render(request, 'robots/index.html')